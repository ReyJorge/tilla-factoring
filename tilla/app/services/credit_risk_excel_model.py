"""
Deterministic invoice financing scoring aligned with invoice_financing_scoring_model_anchor_risk_FINAL.xlsx.

Loads optional workbook (openpyxl, data_only=False) from knowledge_base/credit_risk/scoring_model/.
Embeds conservative defaults when the file is absent or cells are unreadable — production scoring
does not depend on Excel recalculation at runtime.
"""

from __future__ import annotations

import logging
import math
import os
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from app.database import BASE_DIR

logger = logging.getLogger(__name__)

SCORING_MODEL_DIR = BASE_DIR / "knowledge_base" / "credit_risk" / "scoring_model"
DEFAULT_MODEL_FILENAME = "invoice_financing_scoring_model_anchor_risk_FINAL.xlsx"

KNOWN_RATINGS = frozenset({"AAA", "AA", "A", "BBB", "BB", "B", "CCC", "NR"})
GATES_ORDER = {"OK": 0, "MANUAL": 1, "STOP": 2}


def resolve_model_workbook_path() -> Path:
    override = os.getenv("CREDIT_RISK_SCORING_MODEL_PATH", "").strip()
    if override:
        return Path(override).expanduser()
    return SCORING_MODEL_DIR / DEFAULT_MODEL_FILENAME


def _norm_sheet(name: str) -> str:
    s = unicodedata.normalize("NFKD", str(name).strip().lower())
    return "".join(ch for ch in s if not unicodedata.combining(ch)).replace(" ", "")


def _norm_label(val: Any) -> str:
    if val is None:
        return ""
    s = unicodedata.normalize("NFKD", str(val).strip().lower())
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"\s+", " ", s)
    return s


def _to_float(val: Any, default: float | None = None) -> float | None:
    if val is None or val == "":
        return default
    if isinstance(val, bool):
        return float(val)
    if isinstance(val, int | float):
        return float(val)
    s = str(val).strip().replace("\xa0", " ").replace(" ", "").replace(",", ".")
    s = re.sub(r"[^\d.\-]", "", s)
    if not s:
        return default
    try:
        return float(s)
    except ValueError:
        return default


def _parse_date(val: Any) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _worst_gate(*gates: str) -> str:
    cur = "OK"
    for g in gates:
        gg = (g or "OK").strip().upper()
        if gg not in GATES_ORDER:
            gg = "MANUAL"
        if GATES_ORDER[gg] > GATES_ORDER[cur]:
            cur = gg
    return cur


def _reference_sheet_presence(sheetnames: list[str]) -> dict[str, bool]:
    """Which canonical workbook tabs appear (normalized match). Used for observability only."""
    norms = [_norm_sheet(n) for n in sheetnames]
    keys = ("Scoring", "Parametry", "Ciselniky", "Historie", "Schvaleni", "Prehled", "Navod")
    out: dict[str, bool] = {}
    for key in keys:
        nk = _norm_sheet(key)
        out[key] = any(nk == sn or nk in sn or sn in nk for sn in norms)
    return out


@dataclass
class RatingRow:
    rating: str
    score: float
    max_advance_pct: float
    fee_premium_pct: float
    gate: str


@dataclass
class ParsedParams:
    """Numeric parameters — defaults mirror typical workbook intent; Excel overrides when parsed."""

    min_paid_invoices: float = 5.0
    max_avg_delay_days: float = 15.0
    max_late_invoice_share_pct: float = 30.0
    min_relationship_months: float = 12.0
    weight_behavior: float = 0.45
    weight_rating: float = 0.35
    weight_concentration: float = 0.20
    penalty_new_anchor_points: float = 15.0
    penalty_mismatch_points: float = 25.0
    penalty_dispute_points: float = 45.0
    concentration_anchor_portfolio_red_pct: float = 25.0
    concentration_supplier_anchor_red_pct: float = 85.0
    concentration_anchor_portfolio_amber_pct: float = 18.0
    concentration_supplier_anchor_amber_pct: float = 70.0
    base_fee_pct: float = 2.0
    risk_band_a_min: float = 82.0
    risk_band_b_min: float = 68.0
    risk_band_c_min: float = 53.0
    gate_manual_penalty_points: float = 10.0
    gate_stop_penalty_points: float = 38.0


DEFAULT_RATING_TABLE: tuple[RatingRow, ...] = (
    RatingRow("AAA", 100.0, 85.0, 0.35, "OK"),
    RatingRow("AA", 96.0, 82.0, 0.45, "OK"),
    RatingRow("A", 92.0, 78.0, 0.65, "OK"),
    RatingRow("BBB", 84.0, 72.0, 0.95, "OK"),
    RatingRow("BB", 73.0, 65.0, 1.75, "MANUAL"),
    RatingRow("B", 56.0, 52.0, 3.4, "STOP"),
    RatingRow("CCC", 38.0, 42.0, 6.5, "STOP"),
    RatingRow("NR", 46.0, 58.0, 3.1, "MANUAL"),
)


def _merge_rating_catalog(parsed_rows: dict[str, RatingRow]) -> dict[str, RatingRow]:
    out = {r.rating: r for r in DEFAULT_RATING_TABLE}
    out.update(parsed_rows)
    return out


def _param_scan(raw: dict[str, Any]) -> ParsedParams:
    p = ParsedParams()

    def grab(patterns: Iterable[str]) -> Any:
        for k, v in raw.items():
            if any(pat in k for pat in patterns):
                return v
        return None

    if (x := _to_float(grab(("min.pocet", "min pocet", "pocet uhrazen")))) is not None:
        p.min_paid_invoices = x
    if (x := _to_float(grab(("max.prumer", "zpozden")))) is not None:
        p.max_avg_delay_days = x
    if (x := _to_float(grab(("podil pozdn", "pozdnich fakt")))) is not None:
        p.max_late_invoice_share_pct = x
    if (x := _to_float(grab(("min.delka", "min delka", "vztahu")))) is not None:
        p.min_relationship_months = x
    if (x := _to_float(grab(("vah", "behavior")))) is not None and 0 < x <= 1:
        p.weight_behavior = x
    if (x := _to_float(grab(("portfolio", "koncentraci")))) is not None:
        # heuristic: anchor share limit — prefer explicit anchor/portfolio phrases in longer labels
        keys_sorted = sorted(raw.keys(), key=len, reverse=True)
        for kk in keys_sorted:
            if "portfolio" in kk or "portfolia" in kk:
                pv = _to_float(raw[kk])
                if pv is not None:
                    p.concentration_anchor_portfolio_red_pct = pv
                break

    return p


def parse_parametry_sheet(ws: Any) -> dict[str, Any]:
    raw: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=1, max_row=400, values_only=True):
        if not row:
            continue
        label = row[0]
        val = row[1] if len(row) > 1 else None
        nk = _norm_label(label)
        if nk:
            raw[nk] = val
    return raw


def parse_ciselnik_anchor_sheet(ws: Any) -> dict[str, RatingRow]:
    """Expect rating codes in column A — scores as first numeric cells; gate token anywhere."""
    found: dict[str, RatingRow] = {}
    for row in ws.iter_rows(min_row=1, max_row=300, values_only=True):
        if not row or row[0] is None:
            continue
        tok = str(row[0]).strip().upper()
        if tok not in KNOWN_RATINGS:
            continue
        gate_v: str | None = None
        nums: list[float] = []
        for cell in row[1:]:
            if cell is None:
                continue
            gs = str(cell).strip().upper()
            if gs in GATES_ORDER:
                gate_v = gs
                continue
            fv = _to_float(cell)
            if fv is not None:
                nums.append(fv)
        if len(nums) < 2:
            continue
        score_v, adv_v = nums[0], nums[1]
        fee_v = nums[2] if len(nums) > 2 else DEFAULT_RATING_TABLE[0].fee_premium_pct
        if gate_v is None:
            gate_v = "OK"
            if tok in {"BB", "NR"}:
                gate_v = "MANUAL"
            elif tok in {"B", "CCC"}:
                gate_v = "STOP"
        found[tok] = RatingRow(tok, score_v, adv_v, fee_v, gate_v)
    return found


def load_workbook_structures(path: Path | str) -> tuple[ParsedParams, dict[str, RatingRow]]:
    path = Path(path)
    params = ParsedParams()
    catalog = _merge_rating_catalog({})
    if not path.is_file():
        logger.warning("Scoring workbook missing — using embedded defaults: %s", path)
        return params, catalog
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]

        wb = load_workbook(filename=str(path), data_only=False)
    except Exception as exc:
        logger.warning("openpyxl load failed (%s); embedded defaults used.", exc)
        return params, catalog

    def find_sheet(*needles: str) -> str | None:
        cand_norm = [_norm_sheet(x) for x in needles]
        for sn in wb.sheetnames:
            ns = _norm_sheet(sn)
            for cn in cand_norm:
                if cn == ns or cn in ns or ns in cn:
                    return sn
        return None

    pn = find_sheet("Parametry")
    if pn:
        raw = parse_parametry_sheet(wb[pn])
        params = _param_scan(raw)

    cn = find_sheet("Ciselniky", "Číselníky")
    if cn:
        parsed_cat = parse_ciselnik_anchor_sheet(wb[cn])
        catalog = _merge_rating_catalog(parsed_cat)

    logger.info(
        "Loaded scoring workbook sheets=%s catalog_ratings=%s reference_sheets_present=%s",
        list(wb.sheetnames),
        sorted(catalog.keys()),
        _reference_sheet_presence(list(wb.sheetnames)),
    )


@dataclass
class DealInputs:
    anchor: str
    supplier: str
    receivable_status: str
    deal_id: str
    invoice_amount: float
    due_date: date
    data_mismatch: bool
    dispute: bool
    anchor_rating: str
    existing_supplier_exposure: float
    existing_anchor_exposure: float
    total_portfolio_exposure: float
    historical_transactions: list[dict[str, Any]] = field(default_factory=list)
    reference_date: date | None = None


def _historical_derived(
    txns: list[dict[str, Any]], ref: date
) -> tuple[int, float, float, float, float, date | None]:
    amounts: list[float] = []
    delays: list[float] = []
    late_ct = 0
    paid_ct = 0
    earliest: date | None = None
    for raw in txns:
        amt = _to_float(raw.get("invoice_amount") or raw.get("amount"), 0.0) or 0.0
        amounts.append(amt)
        paid_d = _parse_date(raw.get("paid_date"))
        due_d = _parse_date(raw.get("due_date"))
        if paid_d and due_d:
            paid_ct += 1
            delay = max(0.0, float((paid_d - due_d).days))
            delays.append(delay)
            if paid_d > due_d:
                late_ct += 1
        if due_d:
            earliest = due_d if earliest is None else min(earliest, due_d)
        if paid_d:
            earliest = paid_d if earliest is None else min(earliest, paid_d)

    avg_delay = sum(delays) / len(delays) if delays else 0.0
    late_share = (late_ct / paid_ct * 100.0) if paid_ct else 0.0
    avg_inv = sum(amounts) / len(amounts) if amounts else 0.0
    rel_months = 0.0
    if earliest:
        rel_months = max(0.0, (ref - earliest).days / 30.44)
    return paid_ct, avg_delay, late_share, avg_inv, rel_months, earliest


def _lin_score_inverse(actual: float, bad_at: float, good_below: bool = True) -> float:
    """100 when good; linear decay toward 0 past threshold."""
    if bad_at <= 0:
        return 100.0
    if good_below:
        if actual <= bad_at:
            return 100.0
        over = actual - bad_at
        return max(0.0, 100.0 - min(100.0, over / bad_at * 100.0))
    if actual >= bad_at:
        return 100.0
    under = bad_at - actual
    return max(0.0, min(100.0, under / bad_at * 100.0))


def _lin_score_direct(actual: float, need_at: float) -> float:
    if need_at <= 0:
        return 100.0
    if actual >= need_at:
        return 100.0
    return max(0.0, actual / need_at * 100.0)


def compute_deal_scoring(
    deal: DealInputs,
    params: ParsedParams | None = None,
    catalog: dict[str, RatingRow] | None = None,
) -> dict[str, Any]:
    params = params or ParsedParams()
    if catalog is None:
        catalog = _merge_rating_catalog({})

    mp_path = resolve_model_workbook_path()

    ref = deal.reference_date or date.today()
    rating_key = (deal.anchor_rating or "").strip().upper() or "NR"
    if rating_key not in catalog:
        rating_key = "NR"

    paid_ct, avg_delay, late_share, avg_hist_inv, rel_months, _earliest = _historical_derived(
        deal.historical_transactions, ref
    )

    new_anchor = paid_ct < params.min_paid_invoices
    invoice_amt = float(deal.invoice_amount)
    avg_hist_safe = avg_hist_inv if avg_hist_inv > 0 else invoice_amt
    volume_ratio = invoice_amt / avg_hist_safe if avg_hist_safe > 0 else 1.0

    days_to_due = float((deal.due_date - ref).days)
    active_days_to_due = float(max(0, math.ceil(days_to_due)))

    score_history = _lin_score_direct(float(paid_ct), params.min_paid_invoices)
    score_delay = _lin_score_inverse(avg_delay, params.max_avg_delay_days, good_below=True)
    score_volatility = _lin_score_inverse(late_share, params.max_late_invoice_share_pct, good_below=True)
    score_relationship = _lin_score_direct(rel_months, params.min_relationship_months)
    score_volume = max(0.0, min(100.0, 100.0 - max(0.0, (volume_ratio - 1.0) * 35.0)))

    penalty_new = params.penalty_new_anchor_points if new_anchor else 0.0
    penalty_mismatch = params.penalty_mismatch_points if deal.data_mismatch else 0.0
    penalty_dispute = params.penalty_dispute_points if deal.dispute else 0.0

    behavior_parts = [
        score_history,
        score_delay,
        score_volatility,
        score_relationship,
        score_volume,
    ]
    behavior_avg = sum(behavior_parts) / len(behavior_parts)
    behavior_score_raw = max(0.0, min(100.0, behavior_avg - penalty_new - penalty_mismatch - penalty_dispute))

    row = catalog[rating_key]
    anchor_rating_score = row.score
    base_gate = row.gate.upper()

    dispute_gate = "STOP" if deal.dispute else "OK"
    mismatch_gate = "MANUAL" if deal.data_mismatch else "OK"

    supplier_exp = deal.existing_supplier_exposure + invoice_amt
    anchor_exp = deal.existing_anchor_exposure + invoice_amt
    port = deal.total_portfolio_exposure

    supplier_to_anchor_pct = (supplier_exp / anchor_exp * 100.0) if anchor_exp > 0 else 0.0
    anchor_to_portfolio_pct = (anchor_exp / port * 100.0) if port > 0 else 0.0
    supplier_to_portfolio_pct = (supplier_exp / port * 100.0) if port > 0 else 0.0

    concentration_flag = "GREEN"
    if (
        anchor_to_portfolio_pct >= params.concentration_anchor_portfolio_red_pct
        or supplier_to_anchor_pct >= params.concentration_supplier_anchor_red_pct
    ):
        concentration_flag = "RED FLAG"
    elif (
        anchor_to_portfolio_pct >= params.concentration_anchor_portfolio_amber_pct
        or supplier_to_anchor_pct >= params.concentration_supplier_anchor_amber_pct
    ):
        concentration_flag = "AMBER"

    concentration_gate = (
        "MANUAL"
        if concentration_flag == "RED FLAG"
        else ("MANUAL" if concentration_flag == "AMBER" else "OK")
    )

    rating_gate = _worst_gate(base_gate, dispute_gate, mismatch_gate, concentration_gate)

    gate_penalty = 0.0
    if rating_gate == "MANUAL":
        gate_penalty = params.gate_manual_penalty_points
    elif rating_gate == "STOP":
        gate_penalty = params.gate_stop_penalty_points

    if concentration_flag == "GREEN":
        concentration_score = 100.0
    elif concentration_flag == "AMBER":
        concentration_score = 68.0
    else:
        concentration_score = 28.0

    score_behavior = behavior_score_raw
    score_rating = anchor_rating_score
    score_concentration = concentration_score
    rating_gate_penalty = gate_penalty

    total_score = (
        params.weight_behavior * score_behavior
        + params.weight_rating * score_rating
        + params.weight_concentration * score_concentration
        - rating_gate_penalty
    )
    total_score = max(0.0, min(100.0, round(total_score, 2)))

    if total_score >= params.risk_band_a_min:
        risk_band = "A"
    elif total_score >= params.risk_band_b_min:
        risk_band = "B"
    elif total_score >= params.risk_band_c_min:
        risk_band = "C"
    else:
        risk_band = "D"

    advance_cap_behavior = max(48.0, min(row.max_advance_pct, 55.0 + behavior_score_raw * 0.35))
    recommended_advance_pct = round(min(row.max_advance_pct, advance_cap_behavior), 2)
    recommended_fee_pct = round(params.base_fee_pct + row.fee_premium_pct, 3)

    result_code = "RC_OK"
    if rating_gate == "STOP":
        result_code = "RC_STOP"
    elif rating_gate == "MANUAL":
        result_code = "RC_MANUAL"
    if concentration_flag == "RED FLAG":
        result_code = "RC_CONC_RED"

    recommendation_model = "CONDITIONAL — verify gates"
    if rating_gate == "STOP":
        recommendation_model = "REJECT — STOP gate"
    elif rating_gate == "MANUAL":
        recommendation_model = "MANUAL REVIEW REQUIRED"
    elif concentration_flag == "RED FLAG":
        recommendation_model = "MANUAL REVIEW — concentration RED FLAG"
    else:
        recommendation_model = "PROCEED SUBJECT TO POLICY CHECKS"

    approval_parts = []
    if rating_gate == "STOP":
        approval_parts.append("Závěr: ZASTAVIT — nelze automaticky schválit.")
    elif rating_gate == "MANUAL":
        approval_parts.append("Závěr: Vyžaduje se ruční schválení credit officerem.")
    else:
        approval_parts.append("Závěr: Automatický model nevylučuje schválení — doplnit dokumentaci.")

    if concentration_flag == "RED FLAG":
        approval_parts.append("Koncentrace: RED FLAG — řízení výjimky.")
    if deal.dispute:
        approval_parts.append("Spor — stop / výjimka dle politiky.")
    if deal.data_mismatch:
        approval_parts.append("Nesoulad dat — doplnění a kontrola.")

    approval_conclusion = " ".join(approval_parts)
    management_summary = (
        f"Skóre celkem {total_score}, pásmo rizika {risk_band}, ratingová brána {rating_gate}. "
        f"Doporučená záloha {recommended_advance_pct} %, poplatek navýšení +{row.fee_premium_pct} p.b. nad bázi {params.base_fee_pct} %. "
        f"Koncentrace {concentration_flag}."
    )

    out: dict[str, Any] = {
        "anchor": deal.anchor,
        "supplier": deal.supplier,
        "receivable_status": deal.receivable_status,
        "deal_id": deal.deal_id,
        "paid_invoice_count": paid_ct,
        "average_delay_days": round(avg_delay, 4),
        "late_invoice_share": round(late_share, 4),
        "relationship_months": round(rel_months, 4),
        "average_historical_invoice": round(avg_hist_inv, 2),
        "days_to_due": round(days_to_due, 4),
        "active_days_to_due": active_days_to_due,
        "new_anchor": new_anchor,
        "volume_ratio": round(volume_ratio, 6),
        "score_history": round(score_history, 4),
        "score_delay": round(score_delay, 4),
        "score_volatility": round(score_volatility, 4),
        "score_relationship": round(score_relationship, 4),
        "score_volume": round(score_volume, 4),
        "penalty_new": penalty_new,
        "penalty_mismatch": penalty_mismatch,
        "penalty_dispute": penalty_dispute,
        "behavior_score_raw": round(behavior_score_raw, 4),
        "anchor_rating_score": anchor_rating_score,
        "rating_gate": rating_gate,
        "score_rating": score_rating,
        "score_behavior": round(score_behavior, 4),
        "score_concentration": round(score_concentration, 4),
        "rating_gate_penalty": rating_gate_penalty,
        "total_score": total_score,
        "risk_band": risk_band,
        "recommended_advance_pct": recommended_advance_pct,
        "recommended_fee_pct": recommended_fee_pct,
        "recommendation": recommendation_model,
        "supplier_exposure": round(supplier_exp, 2),
        "anchor_exposure": round(anchor_exp, 2),
        "supplier_to_anchor_pct": round(supplier_to_anchor_pct, 4),
        "anchor_to_portfolio_pct": round(anchor_to_portfolio_pct, 4),
        "supplier_to_portfolio_pct": round(supplier_to_portfolio_pct, 4),
        "concentration_score": round(concentration_score, 4),
        "concentration_flag": concentration_flag,
        "result_code": result_code,
        "approval_conclusion": approval_conclusion,
        "management_summary": management_summary,
        "anchor_rating_input": rating_key,
        "model_params_source": mp_path.name if mp_path.is_file() else "embedded_defaults",
    }
    return out


def load_model_bundle() -> tuple[ParsedParams, dict[str, RatingRow]]:
    path = resolve_model_workbook_path()
    return load_workbook_structures(path)
